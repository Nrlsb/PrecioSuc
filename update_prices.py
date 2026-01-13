import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os

# --- Configuration ---
SOURCE_FILE = os.path.join('public', 'Lista001.xlsx')
TARGET_FILE = os.path.join('public', 'ListaDeProductos.xlsx')
TARGET_SHEET_NAME = '001'

# Column names configuration
KEY_COLUMN = 'Cod.Producto'  # Common column in both files
TARGET_PRICE_COLUMN_HEADER = 'Precio Venta' # Column header in Target file to update

# --- Main Script ---

def main():
    print(f"--- Iniciando actualización de precios ---")
    
    # 1. Load Source Data (Lista001.xlsx)
    print(f"Leyendo archivo origen: {SOURCE_FILE}")
    if not os.path.exists(SOURCE_FILE):
        print(f"ERROR: No se encontró el archivo {SOURCE_FILE}")
        return

    try:
        # Read source excel
        df_source = pd.read_excel(SOURCE_FILE)
        print(f"  -> Columnas encontradas en origen: {list(df_source.columns)}")
        
        # Verify key column exists
        if KEY_COLUMN not in df_source.columns:
            print(f"ERROR: La columna clave '{KEY_COLUMN}' no existe en {SOURCE_FILE}")
            return

        # Try to find the price column in source (auto-detect or assume)
        # We look for columns containing 'precio', 'price', 'venta' case-insensitive
        possible_price_cols = [c for c in df_source.columns if 'precio' in str(c).lower() or 'price' in str(c).lower()]
        
        source_price_col = None
        if len(possible_price_cols) == 1:
            source_price_col = possible_price_cols[0]
            print(f"  -> Se detectó columna de precio en origen: '{source_price_col}'")
        elif len(possible_price_cols) > 1:
            # Prefer 'Precio Venta' or 'Precio' if available
            if 'Precio Venta' in possible_price_cols:
                source_price_col = 'Precio Venta'
            elif 'Precio' in possible_price_cols:
                source_price_col = 'Precio'
            else:
                print(f"ADVERTENCIA: Múltiples columnas posibles para precio: {possible_price_cols}")
                source_price_col = possible_price_cols[0] # Take first one
                print(f"  -> Se seleccionó: '{source_price_col}'")
        else:
            print("ERROR: No se detectó ninguna columna de precio en el archivo origen. (Buscando 'precio' o 'price')")
            return

        # Create a dictionary for fast lookup: Code -> Price
        # Clean up codes to strings for consistent matching
        df_source[KEY_COLUMN] = df_source[KEY_COLUMN].astype(str).str.strip()
        price_map = pd.Series(df_source[source_price_col].values, index=df_source[KEY_COLUMN]).to_dict()
        print(f"  -> Se cargaron {len(price_map)} precios del archivo origen.")

    except Exception as e:
        print(f"ERROR al leer archivo origen: {e}")
        return

    # 2. Update Target Data (ListaDeProductos.xlsx)
    print(f"\nLeyendo archivo destino: {TARGET_FILE}")
    if not os.path.exists(TARGET_FILE):
        print(f"ERROR: No se encontró el archivo {TARGET_FILE}")
        return

    try:
        wb = openpyxl.load_workbook(TARGET_FILE)
        if TARGET_SHEET_NAME not in wb.sheetnames:
            print(f"ERROR: La hoja '{TARGET_SHEET_NAME}' no existe en {TARGET_FILE}")
            print(f"  -> Hojas disponibles: {wb.sheetnames}")
            return
        
        ws = wb[TARGET_SHEET_NAME]
        
        # Find header mapping in Target Sheet
        # valid_headers identifies which column index corresponds to 'Cod.Producto' and 'Precio Venta'
        header_row_idx = 1 # Assuming row 1 has headers
        
        col_indices = {}
        for cell in ws[header_row_idx]:
            if cell.value:
                col_indices[str(cell.value).strip()] = cell.column # cell.column is 1-based index (int) or letter? It's integer in recent openpyxl versions usually, let's verify.
                # Actually cell.column is integer in read-only mode or valid index.
                # Let's use cell.column_letter or just cell.column (int)
        
        # openpyxl returns cell.column as int (1, 2, 3...)
        
        # Map back to letter for info
        col_map = {name: idx for idx, cell in enumerate(ws[1], 1) if (name := str(cell.value).strip())}
        
        if KEY_COLUMN not in col_map:
            print(f"ERROR: Columna clave '{KEY_COLUMN}' no encontrada en fila 1 de {TARGET_SHEET_NAME}")
            return
            
        if TARGET_PRICE_COLUMN_HEADER not in col_map:
            print(f"ERROR: Columna destino '{TARGET_PRICE_COLUMN_HEADER}' no encontrada en fila 1 de {TARGET_SHEET_NAME}")
            return
            
        key_col_idx = col_map[KEY_COLUMN]
        target_price_col_idx = col_map[TARGET_PRICE_COLUMN_HEADER]
        
        print(f"  -> Columna '{KEY_COLUMN}' es la numero {key_col_idx}")
        print(f"  -> Columna '{TARGET_PRICE_COLUMN_HEADER}' es la numero {target_price_col_idx}")

        updates_count = 0
        not_found_count = 0
        
        # Iterate rows starting from row 2
        for row in ws.iter_rows(min_row=2):
            # Get code from key column
            # row is a tuple of cells. row[0] is col 1. so indx = key_col_idx - 1
            code_cell = row[key_col_idx - 1]
            code_val = str(code_cell.value).strip() if code_cell.value is not None else ""
            
            if code_val in price_map:
                new_price = price_map[code_val]
                
                # Get price cell
                price_cell = row[target_price_col_idx - 1]
                
                # Update value
                # Check if it's different to avoid useless writes? (optional)
                try:
                    current_val = float(price_cell.value) if price_cell.value is not None else 0
                    if current_val != float(new_price):
                        price_cell.value = new_price
                        updates_count += 1
                except:
                    # If current val is not a number, just overwrite
                    price_cell.value = new_price
                    updates_count += 1
            else:
                not_found_count += 1

        print(f"\nResumen:")
        print(f"  -> Precios actualizados: {updates_count}")
        print(f"  -> Productos sin coincidencia o sin cambio necesario (o no encontrados en origen): {not_found_count}")

        # Save
        output_file = os.path.join('public', 'ListaDeProductos_Updated.xlsx')
        print(f"\nGuardando cambios en: {output_file}")
        wb.save(output_file)
        print("¡Proceso completado con éxito!")
        print(f"Por favor verifica el archivo '{output_file}' y renombralo a '{TARGET_FILE}' si todo está correcto.")

    except Exception as e:
        print(f"ERROR crítico procesando archivo destino: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
